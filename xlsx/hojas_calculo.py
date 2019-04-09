from openpyxl import load_workbook

FILE_PATH = 'Hoja1.xlsx'
HOJA = 'Hoja 1'

workbook =  load_workbook(FILE_PATH,read_only=True)
hoja = workbook[HOJA]

# for row in hoja.iter_rows():
#     print(row[0].value)
#     print(row[1].value)
#     print(row[2].value)
#     print(row[3].value)

for apellido, nombre, edad, email in hoja.iter_rows(min_row=2):
    print(apellido.value)
    print(nombre.value)
    print(edad.value)
    print(email.value)